# imports
# import spotipy.oauth2 as oauth2
import math
from openpyxl import Workbook
from operator import itemgetter

from django.views import View
from django.db.models import Q
from django.urls import reverse
from django.conf import settings
from django.utils import timezone
from django.contrib import messages
from django.contrib.auth import get_user_model
from django.db.models import Avg, Count, Min, Sum
from django.contrib.auth import views as auth_views
from django.utils.decorators import method_decorator
from django.contrib.auth.decorators import login_required
from django.http import HttpResponse, HttpResponseRedirect, FileResponse
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required, permission_required
from django.contrib.auth import authenticate, login, logout, update_session_auth_hash
from django.contrib.admin.views.decorators import staff_member_required, user_passes_test
from django.contrib.auth.forms import UserCreationForm, AuthenticationForm, UserChangeForm, PasswordChangeForm

from accounts import models as account_models
from accounts import forms as account_forms

User = get_user_model()
# End: imports -----------------------------------------------------------------

def getIndexOfTuple(tuplelist, index, value):
    for i, tuple in enumerate(tuplelist):
        if tuple[index] == value:
            return i
    return None



profile_dec = [
    login_required,
]

@method_decorator(profile_dec, name='dispatch')
class ProfileView(View):
    template = "accounts/profile.html"

    def get(self, request, *args, **kwargs):

        return render(request, self.template, {
        })



@method_decorator(profile_dec, name='dispatch')
class EditProfileView(View):
    template = "accounts/edit_profile.html"
    form_class = account_forms.EditUserForm

    def get(self, request, *args, **kwargs):
        form = self.form_class(instance=request.user)
        return render(request, self.template, {'form': form})

    def post(self, request, *args, **kwargs):
        form = self.form_class(data=request.POST, instance=request.user)
        if form.is_valid():
            form.save()
            messages.add_message(request, messages.SUCCESS, f"Profilen din har blitt oppdatert")
            return redirect('accounts:profile')
        else:
            return render(request, self.template, {'form': form})


class SignUpView(View):
    template = "accounts/registration_form.html"
    form_class = account_forms.SignUpForm

    def get(self, request, *args, **kwargs):
        form = self.form_class()
        return render(request, self.template, {'form': form})

    def post(self, request, *args, **kwargs):
        form = self.form_class(request.POST)
        if form.is_valid():
            user = form.save()
            login(request, user)

            try:
                code = form.cleaned_data['code']
                group = account_models.PermissionCode.objects.get(secret=code).group
                user.groups.add(group)
                messages.add_message(request, messages.SUCCESS, f"Med koden '{code}' har du blitt lagt til i avdeling: {group.name}")
            except:
                messages.add_message(request, messages.INFO, f"Koden '{code}' tilsvarer ingen avdeling. Ta kontakt med admin")

            return redirect('home')
        else:
            return render(request, self.template, {'form': form})


@method_decorator(profile_dec, name='dispatch')
class DeleteUserView(View):

    def get(self, request, *args, **kwargs):
        request.user.delete()
        logout(request)
        messages.add_message(request, messages.SUCCESS, f"Brukeren din har blitt slettet fra systemet")
        return redirect('home')


# Should use built in template AuthenticationForm
class LoginView(View):
    template = "accounts/login.html"

    def get(self, request, *args, **kwargs):
        return render(request, self.template)

    def post(self, request, *args, **kwargs):
        email = request.POST['email']
        password = request.POST['password']
        user = authenticate(request, username=email, password=password)
        error = None
        if user is not None:
            login(request, user)
            return redirect('accounts:profile')
        else:
            error = "Feil"

        return render(request, self.template, {'error': error})



@method_decorator(profile_dec, name='dispatch')
class LogoutView(View):

    def get(self, request, *args, **kwargs):
        logout(request)
        return redirect('accounts:login')


@method_decorator(profile_dec, name='dispatch')
class ChangePasswordView(View):
    template = "accounts/change_password.html"
    form_class = account_forms.CustomPasswordChangeForm
    #form_class = PasswordChangeForm

    def get(self, request, *args, **kwargs):
        form = self.form_class(request=request)
        return render(request, self.template, {'form': form})

    def post(self, request, *args, **kwargs):
        form = self.form_class(data=request.POST, request=request)
        if form.is_valid():
            user = form.save()
            update_session_auth_hash(request, user)  # Important!
            return redirect("accounts:profile")
        return render(request, self.template, {'form': form})



export_perms = [
    login_required,
    user_passes_test(lambda u: u.is_superuser, login_url="forbidden")
]

@method_decorator(export_perms, name='dispatch')
class UsersXLSX(View):

    def get(self, request, *args, **kwargs):
        filename = "users.xlsx"

        users = User.objects.all()

        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename={filename}-{timezone.now()}.xlsx'

        workbook = Workbook()

        # Get active worksheet/tab
        worksheet = workbook.active
        worksheet.title = 'Users'

        # Define the titles for columns
        columns = [
            'ID',
            'email',
            'nickname',
            'first_name',
            'last_name',
            'department',
            'staff',
            'superuser',
        ]
        row_num = 1

        # Assign the titles for each cell of the header
        for col_num, column_title in enumerate(columns, 1):
            cell = worksheet.cell(row=row_num, column=col_num)
            cell.value = column_title

        # Iterate through all movies
        for user in users:
            row_num += 1

            # Define the data for each cell in the row
            row = [
                user.id,
                user.email,
                user.nickname,
                user.first_name,
                user.last_name,
                user.department,
                user.is_staff,
                user.is_superuser,
            ]

            # Assign the data for each cell of the row
            for col_num, cell_value in enumerate(row, 1):
                cell = worksheet.cell(row=row_num, column=col_num)
                cell.value = cell_value

        workbook.save(response)

        return response
